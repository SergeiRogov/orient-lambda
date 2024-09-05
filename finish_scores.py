import json
import math
import os
import pandas as pd
from openpyxl.styles import PatternFill, Border, Alignment, Side, Font
from openpyxl import load_workbook


class FinishScores:

    FILE_NAME = 'finish_scores.xlsx'
    MAX_SCORE_A = 100
    MAX_SCORE_B = 75
    MAX_SCORE_C = 50
    PERCENT_OF_EVENTS_CONSIDERED = 0.8

    def __init__(self, year, num_events):
        self.year = year
        self.num_of_events = num_events
        self.num_of_events_considered = math.floor(self.num_of_events * FinishScores.PERCENT_OF_EVENTS_CONSIDERED)

    def add_event_to_overall_table(self, event_title, event_json_filename):

        event_finish_scores = FinishScores._extract_finish_scores(event_json_filename)
        self._add_new_finish_scores(event_finish_scores, event_title)

    @staticmethod
    def _extract_finish_scores(event_json_filename):
        with open(event_json_filename, 'r') as file:
            result_json = json.load(file)

        controls_a = result_json['controls'][0]
        num_of_controls_a = len(controls_a)
        last_control_a = controls_a[-2].split('(')[-1].strip(')')

        controls_b = result_json['controls'][1]
        num_of_controls_b = len(controls_b)
        last_control_b = controls_b[-2].split('(')[-1].strip(')')

        controls_c = result_json['controls'][2]
        num_of_controls_c = len(controls_c)
        last_control_c = controls_c[-2].split('(')[-1].strip(')')

        finish_scores_dict = {}

        best_finish_a = None
        runners_a = result_json['runners'][0]
        for runner in runners_a:
            runner_splits = runner['splits']
            if len(runner_splits) == num_of_controls_a and runner_splits[-1][1] == 1:
                best_finish_a = runner_splits[-1][5]
                break

        best_finish_b = None
        runners_b = result_json['runners'][1]
        for runner in runners_b:
            runner_splits = runner['splits']
            if len(runner_splits) == num_of_controls_b and runner_splits[-1][1] == 1:
                best_finish_b = runner_splits[-1][5]
                break

        best_finish_c = None
        runners_c = result_json['runners'][2]
        for runner in runners_c:
            runner_splits = runner['splits']
            if len(runner_splits) == num_of_controls_c and runner_splits[-1][1] == 1:
                best_finish_c = runner_splits[-1][5]
                break

        # all 3 are the same
        if last_control_a == last_control_b and last_control_b == last_control_c:

            best_finish = min(best_finish_a, best_finish_b, best_finish_c)
            # A Course
            for runner in runners_a:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_a:
                    finish_scores_dict[runner['name']] = round(best_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
            # B Course
            for runner in runners_b:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_b:
                    finish_scores_dict[runner['name']] = round(best_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
            # C Course
            for runner in runners_c:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_c:
                    finish_scores_dict[runner['name']] = round(best_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)

        elif last_control_a == last_control_b:

            best_a_b_finish = min(best_finish_a, best_finish_b)
            # A Course
            for runner in runners_a:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_a:
                    finish_scores_dict[runner['name']] = round(best_a_b_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
            # B Course
            for runner in runners_b:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_b:
                    finish_scores_dict[runner['name']] = round(best_a_b_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
            # C Course
            for runner in runners_c:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_c:
                    finish_scores_dict[runner['name']] = round(best_finish_c / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_C)

        elif last_control_b == last_control_c:

            best_b_c_finish = min(best_finish_b, best_finish_c)
            # A Course
            for runner in runners_a:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_a:
                    finish_scores_dict[runner['name']] = round(best_finish_a / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
            # B Course
            for runner in runners_b:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_b:
                    finish_scores_dict[runner['name']] = round(best_b_c_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_B)
            # C Course
            for runner in runners_c:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_c:
                    finish_scores_dict[runner['name']] = round(best_b_c_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_B)

        elif last_control_a == last_control_c:

            best_a_c_finish = min(best_finish_a, best_finish_c)
            # A Course
            for runner in runners_a:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_a:
                    finish_scores_dict[runner['name']] = round(best_a_c_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
            # B Course
            for runner in runners_b:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_b:
                    finish_scores_dict[runner['name']] = round(best_finish_b / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_B)
            # C Course
            for runner in runners_c:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_c:
                    finish_scores_dict[runner['name']] = round(best_a_c_finish / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
        else:

            # A Course
            for runner in runners_a:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_a:
                    finish_scores_dict[runner['name']] = round(best_finish_a / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_A)
            # B Course
            for runner in runners_b:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_b:
                    finish_scores_dict[runner['name']] = round(best_finish_b / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_B)
            # C Course
            for runner in runners_c:
                runner_splits = runner['splits']
                if len(runner_splits) == num_of_controls_c:
                    finish_scores_dict[runner['name']] = round(best_finish_c / float(
                        runner_splits[-1][5]) * FinishScores.MAX_SCORE_C)

        return finish_scores_dict

    def _get_table(self):
        if not os.path.exists(FinishScores.FILE_NAME):
            with pd.ExcelWriter(FinishScores.FILE_NAME, engine='xlsxwriter') as writer:
                df = pd.DataFrame()
                df.to_excel(writer, sheet_name=self.year, index=False)
        else:
            with pd.ExcelWriter(FinishScores.FILE_NAME, engine='openpyxl', mode='a') as writer:  # 'a' for append mode
                book = writer.book
                if self.year not in book.sheetnames:
                    df = pd.DataFrame()
                    df.to_excel(writer, sheet_name=self.year, index=False)
        df = pd.read_excel(FinishScores.FILE_NAME, sheet_name=self.year)
        return df

    def _add_new_finish_scores(self, event_finish_scores, event_title):

        df = self._get_table()

        event_scores_df = pd.DataFrame.from_dict(event_finish_scores, orient='index', columns=[event_title])
        event_scores_df.index.name = 'Name'
        event_scores_df.reset_index(inplace=True)

        if df.empty:
            df = event_scores_df
        else:
            df = pd.merge(df, event_scores_df, on='Name', how='outer')

        event_columns = [col for col in df.columns if col != 'Name' and col != 'Total' and col != 'Position']

        def _calculate_scores(row):
            scores = [row[col] for col in event_columns if pd.notna(row[col])]
            sorted_scores = sorted(scores, reverse=True)
            top_scores = sorted_scores[:self.num_of_events_considered]
            return sum(top_scores)

        def get_color(score):
            color_ranges = [
                {'min': 0, 'max': 20, 'color': 'f73333'},  # Lowest performance (red)
                {'min': 20, 'max': 30, 'color': 'ff5c23'},
                {'min': 30, 'max': 40, 'color': 'ff7e0c'},
                {'min': 40, 'max': 50, 'color': 'ff9f00'},
                {'min': 50, 'max': 60, 'color': 'ffbe00'},
                {'min': 60, 'max': 70, 'color': 'ffdc00'},
                {'min': 70, 'max': 80, 'color': 'dddf00'},
                {'min': 80, 'max': 90, 'color': 'b7e100'},
                {'min': 90, 'max': 95, 'color': '8ae100'},
                {'min': 95, 'max': 100, 'color': '4be116'},  # Best performance (green)
            ]
            for color_range in color_ranges:
                if color_range['min'] < score <= color_range['max']:
                    return PatternFill(start_color=color_range['color'], end_color=color_range['color'], fill_type="solid")
            return None

        df['Total'] = df.apply(_calculate_scores, axis=1)
        df = df.sort_values(by='Total', ascending=False)
        df['Position'] = df['Total'].rank(method='min', ascending=False).astype(int)
        df = df[['Position'] + ['Name'] + event_columns + ['Total']]

        with pd.ExcelWriter(FinishScores.FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=self.year, index=False)

        book = load_workbook(FinishScores.FILE_NAME)
        sheet = book[self.year]

        name_column_letter = 'B'
        sheet.column_dimensions[name_column_letter].width = 27

        dropped_score_font = Font(strike=True)

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')

        bold_font = Font(bold=True)

        name_column_index = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == 'Name':
                name_column_index = col
                break

        if name_column_index is None:
            raise ValueError("Could not find the 'Name' column in the sheet.")

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
                if row[0].row == 1:
                    cell.alignment = center_alignment
                else:
                    if cell.column == name_column_index:
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment

        for col_index, col in enumerate(event_columns, start=3):
            column_cells = list(sheet.iter_cols(min_col=col_index, max_col=col_index, min_row=2, max_row=sheet.max_row))
            col_cells = [cell for cell in column_cells[0] if cell.value is not None]
            sorted_col_scores = sorted([cell.value for cell in col_cells], reverse=True)

            top_3_scores = []
            for score in sorted_col_scores:
                if score not in top_3_scores:
                    top_3_scores.append(score)
                if len(top_3_scores) == 3:
                    break

            for cell in col_cells:
                if cell.value == 100:
                    cell.font = bold_font

        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=len(event_columns) + 2):
            scores = [cell.value for cell in row]
            sorted_scores = sorted([score for score in scores if score is not None], reverse=True)
            top_scores = set(sorted_scores[:self.num_of_events_considered])
            for cell in row:
                if cell.value is not None:
                    cell.fill = get_color(cell.value)
                    if cell.value not in top_scores:
                        cell.font = dropped_score_font

        book.save(FinishScores.FILE_NAME)


if __name__ == "__main__":

    events_2023 = {
        "Drys": "HTML Kato Drys 22 Jan 23 splits.json",
        "Kouris": "HTML Kouris 26 Feb 23 Splits.json",
        "Pikni": "HTML - Pikni Forest 05 Mar 23 Splits.json",
        "Sia": "Sia Mathiatis 26 Mar 2023 splits.json",
        "Palechori": "HTML - Palechori 07 May 23 splits.json",
        "Olympus": "HTML - Mt Olympus 16 June 2023 splits.json",
        "Troodos": "HTML Troodos 16 Jul 23 splits.json",
        "Piale": "HTML - Piale Pasha 19 Aug 23 splits.json",
        "Kalavasos": "Kalavasos30_09_23splits.json",
        "Souni": "Souni splits as HTML.json",
        "Delikipos": "Splits Championships 2023.json"
    }

    events_2024 = {
        "Lefkara": "Lefkara28-1-24 splits.json",
        "Kouris": "Kouris18-2-24 splits.json",
        "Kalavasos": "21 Apr 24 Kalavasos Splits.json",
        "Palechori": "Palechori 19 May splits.json",
        "Troodos": "Splits Troodos as HTML.json",
        "Olympus": "Mt Olympos 7 Jul 24 Splits as HTML.json",
        "Piale": "Piale Pasha 4 Aug 2024 splits.json"
    }

    jsons_dir = os.path.join(os.path.dirname(__file__), '..', "json's")

    finish_scores_2023 = FinishScores("2023", 11)
    for event_title, event_json_filename in events_2023.items():
        json_file_path = os.path.join(jsons_dir, event_json_filename)
        finish_scores_2023.add_event_to_overall_table(event_title, json_file_path)

    finish_scores_2024 = FinishScores("2024", 10)
    for event_title, event_json_filename in events_2024.items():
        json_file_path = os.path.join(jsons_dir, event_json_filename)
        finish_scores_2024.add_event_to_overall_table(event_title, json_file_path)
